# -*- coding: utf-8 -*-
"""
Created on Fri Mar 17 10:02:58 2023

@author: YOGB
"""
import streamlit as st
import folium
from streamlit_folium import st_folium
from matplotlib import pyplot as plt
import pandas as pd
import openpyxl
import utm
import numpy as np
from scipy.interpolate import griddata
from openpyxl import Workbook
import geojsoncontour
from zipfile import ZipFile
import base64

st.set_page_config(page_title="wibogina-generator", layout="wide")

zipObj = ZipFile("merged_file.zip", "w")

col1, col2 = st.columns([0.5,1.5])

with col1:
    uploaded_files = st.file_uploader("Upload BH/CPT files:", type=["xlsx"], accept_multiple_files=True)
    
    check1 = st.checkbox("Add topography data points", value=False)
    if check1:
        uploaded_topo = st.file_uploader("Upload topography data:", type=["xlsx"], accept_multiple_files=False) 
        topo_zone = st.number_input("UTM zone:", min_value=46, max_value=54)
        topo_hemi = st.selectbox("UTM hemisphere:", options=["North","South"])
    
    if uploaded_files:
        listfile = [i.name for i in uploaded_files]
    
        ID, BH_label, X, Y, Z, water_level, elev_unit, UTM_zone, UTM_NS, city, province, project_code = [],[],[],[],[],[],[],[],[],[],[],[]
        lithnaming, surveyor, date, remark, max_depth = [],[],[],[],[]
        for file in uploaded_files:
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb["general"]
            ID.append(sheet.cell(1,2).value)
            BH_label.append(sheet.cell(2,2).value)
            X.append(sheet.cell(3,2).value)
            Y.append(sheet.cell(4,2).value)
            Z.append(sheet.cell(5,2).value)
            water_level.append(sheet.cell(6,2).value)
            elev_unit.append(sheet.cell(7,2).value)
            UTM_zone.append(sheet.cell(8,2).value)
            UTM_NS.append(sheet.cell(9,2).value)
            city.append(sheet.cell(10,2).value)
            province.append(sheet.cell(11,2).value)
            project_code.append(sheet.cell(12,2).value)
            lithnaming.append(sheet.cell(13,2).value)
            surveyor.append(sheet.cell(14,2).value)
            remark.append(sheet.cell(16,2).value)
            if isinstance(sheet.cell(15,2).value,str):
                date.append(sheet.cell(15,2).value)
            elif isinstance(sheet.cell(15,2).value,int):
                date.append(str(sheet.cell(15,2).value))
            else:
                date.append(str(sheet.cell(15,2).value))
            
            if "spt" in file.name:
                sheet2 = wb['nspt']
                spt_n = sheet2.max_row - 1
                spt_depth = []
                for i in range(spt_n):
                    spt_depth.append(sheet2.cell(i+2,1).value)
                bottom = np.max(spt_depth)
                max_depth.append(bottom)
            elif "cpt" in file.name:
                sheet2 = wb['cpt']
                spt_n = sheet2.max_row - 1
                spt_depth = []
                for i in range(spt_n):
                    spt_depth.append(sheet2.cell(i+2,1).value)
                bottom = np.max(spt_depth)
                max_depth.append(bottom)
            
        data = {'ID':ID,
               'BH_label':BH_label,
               'X_UTM':X,
               'Y_UTM':Y,
               'Z':Z,
               'water_level':water_level,
               'elev_unit':elev_unit,
               'UTM_zone': UTM_zone,
               'UTM_NS': UTM_NS,
               'city':city,
               'provice':province,
               'project_code':project_code,
               'naming':lithnaming,
               'surveyor':surveyor,
               'date':date,
               'remark':remark,
               'max_depth':max_depth}
        
        df = pd.DataFrame(data)
        
        latlon, lat, lon = [],[],[]
        for i in range(len(df)):
            if df["UTM_NS"].iloc[i] == 'N':
                hemis = True
            elif df["UTM_NS"].iloc[i] == 'S':
                hemis = False
            latlon.append(utm.to_latlon(df["X_UTM"].iloc[i], df["Y_UTM"].iloc[i], df["UTM_zone"].iloc[i], northern=hemis))
            lat.append(latlon[i][0])
            lon.append(latlon[i][1])
        
        df['lat']=lat
        df['lon']=lon
        
        latmean = df['lat'].mean()
        lonmean = df['lon'].mean()
        location = [latmean, lonmean]
        m = folium.Map(location=location, zoom_start=16
                        # , tiles="CartoDB positron"
                        )
        
        icons = []
        for i in range(0,len(df)):
            if "spt" in df['ID'].iloc[i]:
                icons.append(folium.Icon(color='black'))
            elif "cpt" in df['ID'].iloc[i]:
                icons.append(folium.Icon(color='red'))
        
        for i in range(0,len(df)):
            folium.Marker([df['lat'].iloc[i], df['lon'].iloc[i]], popup="ID: %s\nLabel: %s"%(df['ID'].iloc[i],df['BH_label'].iloc[i]), icon=icons[i]).add_to(m)
        
        with col2:
            mymap = st_folium(m, height=700, width=1000, returned_objects=[])
        
        # CREATE WIBOGINA INPUT FILE===========================================
        filename = "wibogina_input.xlsx"
        wb_out = Workbook()
        wb_out.remove(worksheet=wb_out['Sheet'])
        
        sheet_bhmgr = wb_out.create_sheet('borehole_manager')
        sheet_cptmgr = wb_out.create_sheet('cpt_manager')
        sheet_litho = wb_out.create_sheet('lithology')
        sheet_nspt = wb_out.create_sheet('nspt')
        sheet_cpt = wb_out.create_sheet('cpt')
        sheet_litholist = wb_out.create_sheet('lithology_list')
        sheet_strati = wb_out.create_sheet('stratigraphy')
        sheet_topo = wb_out.create_sheet('topography')
        sheet_other = wb_out.create_sheet('others')
        sheet_lab = wb_out.create_sheet('labtest')
        
        compilebutton = st.button("Merge", use_container_width=True)
        
        if compilebutton:            
            field_bhmgr = ['No','BH_ID','X','Y','Z','GWL']
            field_cptmgr = ['No','CPT_ID','X','Y','Z','GWL']
            for i in range(len(field_bhmgr)):
                sheet_bhmgr.cell(1,i+1).value = field_bhmgr[i]
                sheet_cptmgr.cell(1,i+1).value = field_cptmgr[i]
                
            for i in range(len(df)):
                if 'spt' in df['ID'].iloc[i]:
                    sheet_bhmgr.cell(i+2,1).value = i+1
                    sheet_bhmgr.cell(i+2,2).value = df['BH_label'].iloc[i]
                    sheet_bhmgr.cell(i+2,3).value = df['X_UTM'].iloc[i]
                    sheet_bhmgr.cell(i+2,4).value = df['Y_UTM'].iloc[i]
                    sheet_bhmgr.cell(i+2,5).value = df['Z'].iloc[i]
                    sheet_bhmgr.cell(i+2,6).value = df['water_level'].iloc[i]
                    sheet_bhmgr.cell(i+2,7).value = df['elev_unit'].iloc[i]
                if 'cpt' in df['ID'].iloc[i]:
                    sheet_cptmgr.cell(i+2,1).value = i+1
                    sheet_cptmgr.cell(i+2,2).value = df['BH_label'].iloc[i]
                    sheet_cptmgr.cell(i+2,3).value = df['X_UTM'].iloc[i]
                    sheet_cptmgr.cell(i+2,4).value = df['Y_UTM'].iloc[i]
                    sheet_cptmgr.cell(i+2,5).value = df['Z'].iloc[i]
                    sheet_cptmgr.cell(i+2,6).value = df['water_level'].iloc[i]
                    sheet_cptmgr.cell(i+2,7).value = df['elev_unit'].iloc[i]
            
            sheet_other.cell(1,1).value = 'shapefile_path'
            sheet_other.cell(2,1).value = 'shapefile_name'
            sheet_other.cell(3,1).value = 'elevation_unit'
            elevunit1 = [i for i in df['elev_unit']]
            elevunit2 = []
            for i in elevunit1:
                if i not in elevunit2:
                    elevunit2.append(i)
            if len(elevunit2) > 1:
                sheet_other.cell(3,2).value = 'non-uniform elevation unit'
            elif len(elevunit2) == 1:
                sheet_other.cell(3,2).value = df['elev_unit'].iloc[0]
            
            for i in range(len(df)):
                if 'spt' in df['ID'].iloc[i]:
                    sheet_litho.cell(1,2*i+1).value = df['BH_label'].iloc[i]
                    sheet_litho.cell(2,2*i+1).value = 'depth'
                    sheet_litho.cell(2,2*i+2).value = 'lithology'
                    sheet_nspt.cell(1,2*i+1).value = df['BH_label'].iloc[i]
                    sheet_nspt.cell(2,2*i+1).value = 'depth'
                    sheet_nspt.cell(2,2*i+2).value = 'nspt'
                    wb_in = openpyxl.load_workbook(uploaded_files[i], data_only=True)
                    sh_lith = wb_in['lithology']
                    sh_nspt = wb_in['nspt']
                    sh_lithtype = wb_in['litholist']
                    for j in range(sh_lith.max_row - 1):
                        sheet_litho.cell(j+3,2*i+1).value = sh_lith.cell(j+2,1).value
                        sheet_litho.cell(j+3,2*i+2).value = sh_lith.cell(j+2,2).value
                    for k in range(sh_nspt.max_row - 1):
                        sheet_nspt.cell(k+3,2*i+1).value = sh_nspt.cell(k+2,1).value
                        sheet_nspt.cell(k+3,2*i+2).value = sh_nspt.cell(k+2,2).value
                    for x in range(sh_lithtype.max_row):
                        sheet_litholist.cell(x+1,1).value = sh_lithtype.cell(x+1,1).value
                        sheet_litholist.cell(x+1,2).value = sh_lithtype.cell(x+1,2).value            
                        sheet_litholist.cell(x+1,3).value = sh_lithtype.cell(x+1,3).value            
                if 'cpt' in df['ID'].iloc[i]:
                    sheet_cpt.cell(1,3*i+1).value = df['BH_label'].iloc[i]
                    sheet_cpt.cell(2,3*i+1).value = 'depth'
                    sheet_cpt.cell(2,3*i+2).value = 'qc'
                    sheet_cpt.cell(2,3*i+3).value = 'fs'
                    wb_in = openpyxl.load_workbook(uploaded_files[i], data_only=True)
                    sh_cpt = wb_in['cpt']
                    for j in range(sh_cpt.max_row - 1):
                        sheet_cpt.cell(j+3,3*i+1).value = sh_cpt.cell(j+2,1).value
                        sheet_cpt.cell(j+3,3*i+2).value = sh_cpt.cell(j+2,2).value
                        sheet_cpt.cell(j+3,3*i+3).value = sh_cpt.cell(j+2,3).value

            # Compiling lab test ------------------------------------------------------
            for i in range(len(df)):
                if "cpt" in df['ID'].iloc[i]:
                    pass
                if "spt" in df['ID'].iloc[i]:
                    wb_in = openpyxl.load_workbook(uploaded_files[i], data_only=True)
                    shin_lab = wb_in["labtest"]
                    if shin_lab.max_row == 0:
                        pass
                    elif shin_lab.max_row == 2:
                        for j in range(shin_lab.max_column):
                            sheet_lab.cell(1,j+1).value = shin_lab.cell(1,j+1).value
                            sheet_lab.cell(2,j+1).value = shin_lab.cell(2,j+1).value
                    elif shin_lab.max_row > 2:
                        for j in range(shin_lab.max_row - 2):
                            start_row = sheet_lab.max_row
                            for k in range(shin_lab.max_column):
                                sheet_lab.cell(1,k+1).value = shin_lab.cell(1,k+1).value
                                sheet_lab.cell(2,k+1).value = shin_lab.cell(2,k+1).value
                                sheet_lab.cell(start_row+j+3,k+1).value = shin_lab.cell(3+j,k+1).value
                    
                    maxrow = sheet_lab.max_row
                    wadah = []
                    for i in range(maxrow):
                       if sheet_lab.cell(i+1,1).value == None:
                            sheet_lab.delete_rows(i+1)

            if check1:    
                if uploaded_topo:
                    wbtopo = openpyxl.load_workbook(uploaded_topo, data_only=True)
                    shtopo = wbtopo[wbtopo.sheetnames[0]]
                    
                    if topo_hemi == "North":
                        hemi = True
                    elif topo_hemi == "South":
                        hemi = False
                    
                    topox, topoy, topoz, topolat, topolon = [],[],[],[],[]
                    for i in range(shtopo.max_row-1):
                        topox.append(shtopo.cell(i+2,1).value)
                        topoy.append(shtopo.cell(i+2,2).value)
                        topoz.append(shtopo.cell(i+2,3).value)
                        sheet_topo.cell(i+2,1).value = topox[i]
                        sheet_topo.cell(i+2,2).value = topoy[i]
                        sheet_topo.cell(i+2,3).value = topoz[i]
                        topolatlon = utm.to_latlon(topox[i], topoy[i], zone_number=topo_zone, northern=hemi)
                        topolat.append(topolatlon[0])
                        topolon.append(topolatlon[1])
            
                    y = np.linspace(min(topolat), max(topolat), 1000)
                    x = np.linspace(min(topolon), max(topolon), 1000)
                    
                    X,Y = np.meshgrid(x, y)
                    Z = griddata((topolon, topolat), topoz, (X, Y), method='cubic')
            
                    contour = plt.contour(X,Y,Z)
                    kontur = geojsoncontour.contour_to_geojson(contour=contour,
                                                                ndigits=5)
                    
                    folium.GeoJson(kontur).add_to(m)
                    folium.LayerControl().add_to(m)
    
            wb_out.save(filename=filename)
            zipObj.write(filename)
            zipObj.close()
            
            ZipfileDotZip = "merged_file.zip"
            
            with open(ZipfileDotZip, "rb") as f:
                bytes = f.read()
                b64 = base64.b64encode(bytes).decode()
                href = f"<a href=\"data:file/zip;base64,{b64}\" download='{ZipfileDotZip}.zip'>\
                    Download ZIP\
                </a>"
            st.markdown(href, unsafe_allow_html=True)
        

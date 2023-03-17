# -*- coding: utf-8 -*-
"""
Created on Fri Mar 17 09:20:39 2023

@author: YOGB
"""
import streamlit as st
import pandas as pd
import openpyxl
from zipfile import ZipFile
import base64

st.set_page_config(page_title="wibogina-generator", layout="centered")

# INPUT =======================================================================

col1, col2 = st.columns(2)
with col1:
    input_filename = st.file_uploader("Upload WIBOGINA file:", type=["xlsx"], accept_multiple_files=False)
    author = st.selectbox("Author:", options=["yogb","popm3","sary","harn5"])
    start_spt_label = st.number_input("Starting number for SPT data entry:",step=1)
    start_cpt_label = st.number_input("Starting number for CPT data entry:",step=1)
    utmz = st.number_input("UTM zone:", min_value=46, max_value=54)
    utmh = st.selectbox("UTM hemisphere:", options=["N","S"])
UTM = [utmz, utmh]
with col2:
    city = st.text_input("City:")
    province = st.text_input("Province:")
    project_code = st.number_input("Project code:",step=1)
    lithonaming = st.text_input("Lithology naming:")
    surveyor = st.text_input("Surveyor:")
    date = st.date_input("Survey date:")
    date = str(date)
    remark = st.text_area("Remarks:")

# LOAD WORKBOOK ===============================================================
if input_filename:
    wb_in = openpyxl.load_workbook(input_filename, data_only=True)
    shin_nspt = wb_in["nspt"]
    shin_lith = wb_in["lithology"]
    shin_cpt = wb_in["cpt"]
    # shin_lithlist = wb_in["lithology_list"]
    shin_other = wb_in["others"]
    
    # FIELD NAMES 
    
    field = ["ID",
    "BH_label",
    "X",
    "Y",
    "Z",
    "water_level",
    "elev_unit",
    "UTM_zone",
    "UTM_NS",
    "city",
    "province",
    "project_code",
    "lithology_naming",
    "surveyor",
    "date",
    "remark"
    ]
    
    
    bhmgr = pd.read_excel(input_filename, sheet_name="borehole_manager")
    litholist = pd.read_excel(input_filename, sheet_name="lithology_list")
    cptmgr = pd.read_excel(input_filename, sheet_name="cpt_manager")
    
    zipObj = ZipFile("extracted_files.zip", "w")
    
    if bhmgr.empty:
        pass
    else:
        BH_ID = [i for i in bhmgr["BH_ID"]]
        BH_X = [i for i in bhmgr["X"]]
        BH_Y = [i for i in bhmgr["Y"]]
        BH_Z = [i for i in bhmgr["Z"]]
        BH_GWL = [i for i in bhmgr["GWL"]]
        litho_ID = [i for i in litholist["lithoID"]]
        litho_col = [i for i in litholist["color"]]
        litho_hat = [i for i in litholist["hatch"]]
        for i in range(len(BH_ID)):
            ID = author + "_spt_%05d"%(start_spt_label+i)
            wbspt = openpyxl.Workbook()
            wbspt.remove(worksheet=wbspt['Sheet'])
            wbspt_gen = wbspt.create_sheet("general")
            wbspt_spt = wbspt.create_sheet("nspt")
            wbspt_lith = wbspt.create_sheet("lithology")
            wbspt_lithlist = wbspt.create_sheet("litholist")
            wbspt_lab = wbspt.create_sheet("labtest")
            
            # GENERAL--------------------------------------------------------------
            isian = [ID,
                     BH_ID[i],
                     BH_X[i],
                     BH_Y[i],
                     BH_Z[i],
                     BH_GWL[i],
                     shin_other.cell(3,2).value,
                     UTM[0],
                     UTM[1],
                     city,
                     province,
                     project_code,
                     lithonaming,
                     surveyor,
                     date,
                     remark
                     ]
            
            for j in range(len(field)):
                wbspt_gen.cell(j+1,1).value = field[j]
                wbspt_gen.cell(j+1,2).value = isian[j]
            
            # NSPT ----------------------------------------------------------------
            depth, nspt = [],[]
            for j in range(50):
                depth.append(shin_nspt.cell(j+3,2*i+1).value)
                nspt.append(shin_nspt.cell(j+3,2*i+2).value)
            
            depth = [j for j in depth if j != None]
            nspt = [j for j in nspt if j != None]
            
            for j in range(len(nspt)):
                wbspt_spt.cell(j+2,1).value = depth[j]
                wbspt_spt.cell(j+2,2).value = nspt[j]
            
            wbspt_spt.cell(1,1).value = "depth"
            wbspt_spt.cell(1,2).value = "nspt"
            
            # Lithology -----------------------------------------------------------
            depthl, lith = [],[]
            for j in range(50):
                depthl.append(shin_lith.cell(j+3,2*i+1).value)
                lith.append(shin_lith.cell(j+3,2*i+2).value)
            
            depthl = [j for j in depthl if j != None]
            lith = [j for j in lith if j != None]
            
            for j in range(len(lith)):
                wbspt_lith.cell(j+2,1).value = depthl[j]
                wbspt_lith.cell(j+2,2).value = lith[j]
            
            wbspt_lith.cell(1,1).value = "depth"
            wbspt_lith.cell(1,2).value = "lithology"
            
            # Lithology list -----------------------------------------------------------
            for j in range(len(litholist)):
                wbspt_lithlist.cell(j+2,1).value = litho_ID[j]
                wbspt_lithlist.cell(j+2,2).value = litho_col[j]
                wbspt_lithlist.cell(j+2,3).value = litho_hat[j]
            
            wbspt_lithlist.cell(1,1).value = "lithoID"
            wbspt_lithlist.cell(1,2).value = "color"
            wbspt_lithlist.cell(1,3).value = "hatch"
            
            # SAVE ----------------------------------------------------------------
            wbspt.save(ID+".xlsx")
            zipObj.write(ID+".xlsx")
    
    if cptmgr.empty:
        pass
    else:
        CPT_ID = [i for i in cptmgr["CPT_ID"]]
        CPT_X = [i for i in cptmgr["X"]]
        CPT_Y = [i for i in cptmgr["Y"]]
        CPT_Z = [i for i in cptmgr["Z"]]
        CPT_GWL = [i for i in cptmgr["GWL"]]
        
        for i in range(len(CPT_ID)):
            ID = author + "_cpt_%05d"%(start_cpt_label+i)
            wbcpt = openpyxl.Workbook()
            wbcpt.remove(worksheet=wbcpt['Sheet'])
            wbcpt_gen = wbcpt.create_sheet("general")
            wbcpt_cpt = wbcpt.create_sheet("cpt")
            
            # GENERAL--------------------------------------------------------------
            isian = [ID,
                     CPT_ID[i],
                     CPT_X[i],
                     CPT_Y[i],
                     CPT_Z[i],
                     CPT_GWL[i],
                     shin_other.cell(3,2).value,
                     UTM[0],
                     UTM[1],
                     city,
                     province,
                     project_code,
                     "Robertson 2010",
                     surveyor,
                     date,
                     remark
                     ]
            
            for j in range(len(field)):
                wbcpt_gen.cell(j+1,1).value = field[j]
                wbcpt_gen.cell(j+1,2).value = isian[j]
            
            # CPT ----------------------------------------------------------------
            depth, qc, fs = [],[],[]
            for j in range(2500):
                depth.append(shin_cpt.cell(j+3,3*i+1).value)
                qc.append(shin_cpt.cell(j+3,3*i+2).value)
                fs.append(shin_cpt.cell(j+3,3*i+3).value)
                
            depth = [j for j in depth if j != None]
            qc = [j for j in qc if j != None]
            fs = [j for j in fs if j != None]
                    
            for j in range(len(qc)):
                wbcpt_cpt.cell(j+2,1).value = depth[j]
                wbcpt_cpt.cell(j+2,2).value = qc[j]
                wbcpt_cpt.cell(j+2,3).value = fs[j]
            
            wbcpt_cpt.cell(1,1).value = "depth_m"
            wbcpt_cpt.cell(1,2).value = "qc_MPa"
            wbcpt_cpt.cell(1,3).value = "fs_MPa"
            
            # SAVE ----------------------------------------------------------------
            wbcpt.save(ID+".xlsx")
            zipObj.write(ID+".xlsx")
    
    zipObj.close()
    
    ZipfileDotZip = "extracted_files.zip"

    with open(ZipfileDotZip, "rb") as f:
        bytes = f.read()
        b64 = base64.b64encode(bytes).decode()
        href = f"<a href=\"data:file/zip;base64,{b64}\" download='{ZipfileDotZip}.zip'>\
            Download ZIP\
        </a>"
    st.sidebar.markdown(href, unsafe_allow_html=True)